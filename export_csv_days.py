#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# -------------------------------------------------------------------
# Date: 2022/7/4
# Author:   Shaun
# 檔案功能描述: 從資料庫收集OEE資料，並轉出CSV。
# 讀取machine_config.xlsx 機台名稱。
# -------------------------------------------------------------------
import time
import threading
import pandas as pd
import sqlalchemy as sqla
from datetime import timedelta
from datetime import datetime
from copy import deepcopy
from openpyxl.styles import Font
from openpyxl.chart.text import RichText
from openpyxl.chart.plotarea import DataTable
from openpyxl.drawing.text import  RichTextProperties,Paragraph,ParagraphProperties, CharacterProperties
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
	BarChart,
	LineChart,
	label
)

import sys,os

class DB_connect:
	# database setting
	__engine = None
	day_min = None
	day_max = None

	def __init__(self,min,max):
		global NOW
		with open ("dbconfig.txt", "r") as dbconfig:
			data=dbconfig.readlines()
		self.__engine=sqla.create_engine('mysql+mysqlconnector://'+data[0])
		self.day_min=min
		self.day_max=max
		print(self.day_min,self.day_max)
		
	def read_from_oee(self,name):
		t1=time.time()

		#select ID between today
		sql = "select a.d as date,b.name, b.OEE, b.Big_A, b.Availability, b.Performance, b.Quality, b.speed, b.Production, b.load_time,\
			 b.total_time, b.standard_pcs, b.actual_pcs, b.ttr, b.ft, b.mtbf, b.mttr \
 			FROM \
			(SELECT max(id) as id, date(date_add(date,interval -8 hour)) as d FROM `oee` WHERE name='" + name + "' and date between '" + str(self.day_min) + "' and '" + str(self.day_max) + "'" \
			 + " GROUP by d ORDER BY `id` DESC) a left JOIN \
			(select * FROM oee) b on a.id=b.id order by date asc;" 
		
		seldf = pd.read_sql_query(sql, self.__engine)
		print('read database cost %f sec' %(time.time()-t1))
		return seldf
		
	def close(self):
		self.__engine.dispose()

	
if __name__ == '__main__':
	# dir=r'\\Nas\uta-share\UTA資料庫\UTA共用區\Q-專案執行\改善\(G組) MES\表單自動化' + '\\'
	dir='/home/uta_iot/excel_output/'
	if len(sys.argv) ==2:
		min,max=sys.argv[1].split("-")
		min=datetime.strptime(min.strip(), "%Y%m%d").date() + timedelta(hours=8)
		max=datetime.strptime(max.strip(), "%Y%m%d").date() + timedelta(hours=8)
		path=dir+sys.argv[1]+'_excel_output.xlsx'
		
	else:
		print("輸入格式錯誤")
		exit()
	allst=time.time()
	# read from machine config
	machinedf=pd.read_excel("machine_config.xlsx")
	machinedf.columns=['name','host','parts','parts_type','status','status_type','UID','UID_type','WID','WID_type','option1','op_type1','option2','op_type2']
	machine = machinedf['name'].values.tolist()
	print(machine)
	print(time.time()-allst)
	if os.path.exists(path):
		os.remove(path)
	wb = openpyxl.Workbook()
	# wb.remove_sheet(wb['Sheet'])
	del wb['Sheet']
	for i in range(len(machine)):
		name=machine[i].lower()	
		wb.create_sheet(title=name)
	wb.save(path)
	writer = pd.ExcelWriter(engine='openpyxl', path=path, mode='a',if_sheet_exists='replace')
	# wb = writer.book
	for i in range(len(machine)):
		conn = DB_connect(min,max)
		name=machine[i].lower()
		seldf = conn.read_from_oee(name)
		print(seldf)
		# if seldf is not None:
		if not seldf.empty:
			mask=['Production', 'load_time', 'total_time', 'ttr', 'mtbf', 'mttr']
			seldf[mask]=seldf[mask].apply(lambda _:pd.to_timedelta(_, unit='s').astype(str).str.replace('0 days ',''))
			
			if not os.path.exists(path):
				pd.DataFrame({}).to_excel(path,sheet_name=name)

			# with pd.ExcelWriter(engine='openpyxl', path=path, mode='a',if_sheet_exists='replace') as writer:
			seldf.to_excel(writer,sheet_name=name,float_format = "%0.2f")
			# wb = writer.book
			ws= writer.sheets[name]
			for row in ws['A:R']:
				for cell in row:
					cell.font = Font(size=16)
			for j in range(1, ws.max_column+1):
				ws.column_dimensions[get_column_letter(j)].bestFit = True
				ws.column_dimensions[get_column_letter(j)].auto_size = True
				if j==2:
					ws.column_dimensions[get_column_letter(j)].width = 20
				elif j in (6,7,13):
					ws.column_dimensions[get_column_letter(j)].width = 16
			
			last=seldf.shape[0]+1
			for row in range(2,last+1):
				ws[f'D{row}'].number_format ='0.00"%"'
				ws[f'E{row}'].number_format ='0.00"%"'
				ws[f'F{row}'].number_format ='0.00"%"'
				ws[f'G{row}'].number_format ='0.00"%"'
			data1 = Reference(ws, min_col=4, min_row=1, max_col=8, max_row=last)
			# data2 = Reference(ws, min_col=6, min_row=last+1, max_col=6, max_row=last)
			titles = Reference(ws, min_col=2, min_row=2, max_row=last)
			chart1 = BarChart()
			chart1.title = name + " OEE Bar Chart"
			chart1.add_data(data=data1, titles_from_data=True)
			chart1.set_categories(titles)
			chart1.y_axis.title = '百分比'
			chart1.y_axis.numFmt='0"%"'
			# chart1.x_axis.title = '日期'
			chart1.legend=None
			pp=ParagraphProperties(defRPr=CharacterProperties(sz=1400))
			rich_text = RichText(bodyPr=RichTextProperties(anchor="ctr",anchorCtr="1",rot="0",
				spcFirstLastPara="1",vertOverflow="ellipsis",wrap="square"),
				p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(sz=1400))])
			rich_text_T=deepcopy(rich_text)
			rich_text_T.bodyPr.rot="-5400000"
			chart1.dataLabels = label.DataLabelList()
			chart1.dataLabels.showVal = True
			chart1.dataLabels.txPr = rich_text_T
			chart1.plot_area.dTable = DataTable()
			chart1.plot_area.dTable.showHorzBorder = True
			chart1.plot_area.dTable.showVertBorder = True
			chart1.plot_area.dTable.showOutline = True
			chart1.plot_area.dTable.showKeys = True
			chart1.y_axis.txPr = rich_text
			chart1.y_axis.title.tx.rich.p[0].pPr = pp
			chart1.plot_area.dTable.txPr = rich_text
			chart1.style = 26
			chart1.height = 15
			chart1.width = 45
			last=last+2
			ws.add_chart(chart1, "A"+str(last))
			ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
			ws.page_setup.paperSize = ws.PAPERSIZE_A3
			ws.sheet_properties.pageSetUpPr.fitToPage = True
			ws.print_options.headings=True
	writer.save()

	conn.close()
	alled = time.time()
	# 列印結果
	print("loop cost %f sec" % (alled - allst))  # 會自動做進位
	



