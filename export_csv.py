#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# -------------------------------------------------------------------
# Date: 2021/11/08
# Author:   Shaun
# 檔案功能描述: 從資料庫收集機台生產資料，計算OEE資料，並轉出CSV。
# DB_connect: 使用class 操作資料庫讀取、寫入及資料庫資料減量。
# OEE_Class: 使用class 計算OEE相關數據。
# 讀取machine_config.xlsx 機台名稱。
#--------------------------------------------------------------------
# schedule 定義
# 早班		08:00-17:00 標準工時 8h   除外工時 1h
# 中班		17:00-23:59 標準工時 6.5h 除外工時 0.5h
# 夜班-隔日	00:00-08:00 標準工時 8h
#
# 除外工時 定義
# 除外工時需做加班檢查
# 中午休息時間	12:00-13:00
# 晚上休息時間	17:00-17:30 累計至中班
# 
# 每日計畫行程自動判斷 定義
# 每日早班計畫 預設	08:00-17:00
# 每日中班計畫 預設	17:30-18:30 每小時累計至 23:59
# 每日晚班計畫 預設	00:00-01:00 每小時累計至 08:00
# 
# 
# -------------------------------------------------------------------
# OEE公式定義
# OEE,標準產量,實際產量  先分班計算,再累計計算(由SQL 語法於dashboard累計)
#1.OEE
#公式:OEE=A*P*Q=(實際機時/標準機時)*(實際產能/標準產能)*1
#A=稼働率  P=產能效率   Q=良率
#A=機台有效工作時間(0:0~now)/機台設定之作業工時(0:0~now) 
#P=機台成品數/(機台開機時間*每小時標準產能)

#2.FTE
#公式:以機器FTE(天)為標準，計算每位操作員的個人FTE。

#3.設備使用率
#公式:設備使用率=實際機時/(製程機台數*24小時)
# -------------------------------------------------------------------
import time
import threading
import re
import math
import pandas as pd
import sqlalchemy as sqla
from datetime import timedelta
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.chart.plotarea import DataTable
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import  RichTextProperties,Paragraph,ParagraphProperties, CharacterProperties
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
	BarChart,
	LineChart,
	label
)

# import jpype
# import asposecells
# jpype.startJVM()
# from asposecells.api import (
# 	Workbook,
# 	PdfSaveOptions
# ) 

import sys,os
VALUE_OF_PLAN_ST =1
VALUE_OF_PLAN_ED =2
DAY_START= datetime.min.time().replace(hour=8)
CHK_TOLERANCE = timedelta(minutes=5)
NOW=datetime.today().date()

# test_date = datetime(2022,4,19,2,25,0)
# NOW=test_date

class OEE_Class:
	machine = ''
	workid = None
	work_time = None
	A = 0
	a = 0
	P = 0
	Q = 0
	OEE = 0
	actual_pcs = 0 #今日生產數量
	total_time = timedelta(days=0) #總時間 - 非工作時間
	load_time = timedelta(days=0) #負荷時間 (扣除休息時間)
	nomal_time = timedelta(days=0) #淨稼動時間 (正常運轉時間)
	capacity = 0 #pcs / hour 標準產能
	standard_pcs = 0 #今日標準產量
	speed = 0 #機台生產速度

	# alarm
	nomal_min = 0
	nomal_max = 0
	nomal_avg = 0
	alarm_min = 0
	larm_max = 0
	alarm_avg = 0
	# alarm top 5
	piedf = None
	now = None
	


	def __init__(self,seldf,name,rest,schdf):
		t1=time.time()
		global NOW
		self.now=NOW
		self.machine = name
		# 2^32溢位問題排除
		# if seldf.parts.max() > 2**32-10**7:
		# 	over_max=seldf[seldf.parts>2**31].max().fillna(0)
		# 	over_min=seldf[seldf.parts>2**31].min().fillna(0)
		# 	below_max=seldf[seldf.parts<=2**31].max().fillna(0)
		# 	below_min=seldf[seldf.parts<=2**31].min().fillna(0)
		# 	self.actual_pcs = over_max-over_min + below_max - below_min
		# else:
		# 	self.actual_pcs = seldf["parts"].max()-seldf["parts"].min() # 2^32溢位問題須排除
		
		morning=schdf[schdf.raw_by=='morning'].date.min()
		last = schdf[schdf.raw_by=='pre_last'].date.max()

		# seldf not empty in main
		st_dt = morning
		end_dt=last if datetime.today() > last else datetime.today()
		

		# add st_dt,end_dt at top,last to more accurate
		seldf=seldf.drop(seldf[seldf.date>last].index)
		seldf.loc[-1]={'date':st_dt}
		seldf.index=seldf.index+1
		seldf = seldf.sort_index().fillna(method="bfill")
		seldf = pd.concat([seldf,pd.Series({'date':end_dt}).to_frame(len(seldf)).T],axis=0).fillna(method="ffill")
		seldf['flag']=seldf['value']==seldf['value'].shift(1)
		seldf.loc[seldf.tail(1).index,'flag']=False
		seldf=seldf.drop(seldf[seldf.flag==True].index)
		print(seldf)

		work_time=pd.DataFrame({},columns=['pre_time','time','work_id', 'status', 'during'])
		work_time['pre_time']=seldf['date']
		work_time['time']=seldf['date'].shift(-1)
		work_time['work_id']=seldf['work_order_id']
		work_time['status']=seldf['value']
		work_time['during']=seldf['date'].shift(-1)-seldf['date']
		
		work_time['parts']=seldf['parts']
		work_time['parts']=work_time['parts'].shift(-1)-work_time['parts']
		work_time['MTBF']=work_time.loc[work_time['status']==1].during.fillna(pd.Timedelta(seconds=0))
		work_time['MTTR']=work_time.loc[(work_time['status']!=1) & (work_time['status']<500)].during.fillna(pd.Timedelta(seconds=0))


		# work_time[['MTBF','MTTR']]=work_time[['MTBF','MTTR']].fillna(pd.Timedelta(seconds=0))
		# work_time=work_time.fillna(pd.Timedelta(seconds=0))
		#work_time.to_csv(name + ".csv")
		print(work_time)
		print('work_time----------------------------------------------------------')
		last_id=work_time.tail(1).index # is a list
		work_time=work_time.drop(last_id)

		# pd.set_option('display.max_rows', work_time.shape[0]+1)
		print(work_time)
		self.nomal_time=work_time['during'].loc[work_time['status']==1].agg('sum')
		self.work_time = work_time
		print('nomal_time ,total_time----------------------------------------------------------')
		print(work_time.groupby(['status'])['during'].agg('sum').reset_index())
		print('during_sum',work_time.during.agg('sum'))
		mask=(work_time['status'] != 1 ) & (work_time['status'] < 500 )
		print('alarm_sum',work_time.loc[mask].during.agg('sum'))
		rest=work_time[work_time.status>=500].during.agg('sum')
		self.actual_pcs=work_time[work_time.parts>0].parts.agg('sum')
		self.total_time = end_dt-st_dt-rest
		self.load_time = end_dt-st_dt-rest
		print('rest',rest)
		print(self.nomal_time , self.total_time)
		print('test st ed rest---------------------------------------')
		print(st_dt,end_dt)
		print('last',last)
		spdf=pd.DataFrame({},columns=['parts', 'during'])
		print('speed-----------------------------------------------------')
		print(self.work_time[self.work_time['status']==1])
		if not self.work_time[self.work_time['status']==1].empty:
			spdf['parts']=self.work_time[self.work_time['status']==1]['parts']
			spdf['during']=self.work_time[self.work_time['status']==1].during.dt.total_seconds()
			spdf['speed']=spdf.apply(lambda x: round( x['parts'] / x['during'] *60,ndigits=0) if x['during']!=0 else 0 , axis=1)
			spdf=spdf[spdf.parts!=0]
			print(spdf)
			# self.speed=max(spdf['speed'])
			self.speed= round(spdf['parts'].sum() / spdf['during'].sum() * 60,ndigits=0)
		else:
			self.speed=0
		
		
		
		self.__alarm_analyze()

		# self.work_time=work_time.groupby(['work_id'])['during'].agg('sum').reset_index()
		# #print(self.work_time['work_id'].astype(int).values.tolist())
		
		# if len(self.work_time)>0:
		# 	self.workid = '(' + str(self.work_time['work_id'].astype(int).values[0]) + ')' if len(self.work_time) <2 \
		# 		else tuple(self.work_time['work_id'].astype(int).values.tolist())
		
		#print(self.workid)
		# print(self.work_time)
		
		

		
	# 2022/03/25 改變算法
	# def calc_standrad(self,workdf):
		# standard=0
		# during = timedelta(days=0)
		# ahour = timedelta(hours=1)
		# # print(self.work_time)
		# for i in range(self.work_time.shape[0]):
			# work_id=self.work_time.iloc[i,0].astype(int)
			# during=self.work_time.iloc[i,1]
			# cap=workdf.loc[work_id,'standard_CAP'].astype(int)
			# standard+=during/ahour*cap
			# # print(standard,cap,during.total_seconds(),ahour.total_seconds())
			# #print(work_id,psc,during,int(standard))
			# #print()
		# self.standard_pcs=standard
		
	# 2022/03/25 改用機台標準速度計算
	# class DB_connect.read_capacity 設變
	# def calc_standrad(self,workdf):
	def calc_standrad(self,cap):
		standard=0
		during = self.nomal_time
		# print(during)
		ahour = timedelta(hours=1)
		# print(self.work_time)
		# print(cap)
		standard=round(during/ahour*cap,0)
		self.standard_pcs=standard
		

	def calc_OEE(self):
		print('---------------------------------------------------')
		print('正常啟動:',self.nomal_time)
		print('負荷時間:',self.load_time)
		print('總時間:',self.total_time)
		self.A=float(self.load_time/self.total_time*100) if self.total_time!=timedelta(0) else 0
		self.a=float(self.nomal_time/self.load_time*100) if self.load_time!=timedelta(0) else 0
		print("負荷時間率",self.A,"%")
		print("稼動率",self.a,"%")
		print("實際數量:",self.actual_pcs)
		print("計畫數量:",self.standard_pcs)
		self.P=float(self.actual_pcs/self.standard_pcs*100) if self.standard_pcs!=0 else 0
		print("產能效率:",self.P,"%")
		self.Q=float(1)
		self.OEE=float(self.a*self.P*self.Q/100)
		print("OEE:",self.OEE,"%")
		print('機台速度',self.speed, 'sec/pcs')
		
		ls=[[self.now, self.machine, self.OEE, self.A, self.a, self.P, self.Q, self.nomal_min, self.nomal_max, self.nomal_avg,
			self.alarm_min, self.alarm_max, self.alarm_avg,self.speed]]
		#print(self.nomal_min, self.nomal_max, self.nomal_avg, self.alarm_min, self.alarm_max, self.alarm_avg)
		oeedf=pd.DataFrame(ls,columns=['date', 'name', 'OEE', 'Big_A', 'Availability', 'Performance', 'Quality',
									   'nomal_min', 'nomal_max', 'nomal_avg', 'alarm_min', 'alarm_max', 'alarm_avg','speed'])
		oeedf=oeedf.fillna(0)
		oeedf['date']=self.now
		oeedf['Production']=self.nomal_time.total_seconds()
		oeedf['load_time']=self.load_time.total_seconds()
		oeedf['total_time']=self.total_time.total_seconds()
		oeedf['standard_pcs']=self.standard_pcs
		oeedf['actual_pcs']=self.actual_pcs
		oeedf['ttr']=self.piedf[(self.piedf.status!=1) & (self.piedf.status < 500)].during.sum()
		oeedf['ft']=self.piedf[(self.piedf.status!=1) & (self.piedf.status < 500)].times.sum()


		if oeedf.loc[0,'ft']!=0:
			oeedf['mtbf']=round((oeedf.total_time-oeedf.ttr)/oeedf.ft,0)
			oeedf['mttr']=round(oeedf.ttr/oeedf.ft)
		else:
			oeedf['mtbf']=0
			oeedf['mttr']=0
		# print(oeedf.loc[0,['ft','ttr','Production','total_time','mtbf','mttr']])
		return oeedf

	def __alarm_analyze(self):
		#work_time=work_time[work_time['during']!=timedelta(days=0)]
		self.nomal_min=self.work_time[self.work_time['status']==1].during.min().total_seconds()
		self.nomal_max=self.work_time[self.work_time['status']==1].during.max().total_seconds()
		self.nomal_avg=self.work_time[self.work_time['status']==1].during.mean().total_seconds()

		self.alarm_min=self.work_time[self.work_time['status']!=1].during.min().total_seconds()
		self.alarm_max=self.work_time[self.work_time['status']!=1].during.max().total_seconds()
		self.alarm_avg=self.work_time[self.work_time['status']!=1].during.mean().total_seconds()
		self.piedf=self.work_time.groupby(['status'])['during'].agg(['sum','count']).reset_index()
		#piedf.columns=['during']
		#piedf['times']=self.work_time.groupby(['status'])
		self.piedf.columns=['status','during','times']
		self.piedf['during']=self.piedf['during'].dt.total_seconds()
		self.piedf.insert(0,'date',self.now)
		self.piedf.insert(0,'name',self.machine)
		print(self.piedf)



class DB_connect:
	# database setting
	__engine = None
	today_min = None
	today_max = None
	today_now = None

	def __init__(self):
		global NOW
		with open ("dbconfig.txt", "r") as dbconfig:
			data=dbconfig.readlines()
		self.__engine=sqla.create_engine('mysql+mysqlconnector://'+data[0])
		self.today_min=datetime.combine(NOW,DAY_START)
		self.today_max=self.today_min + timedelta(days=1)
		print(self.today_min,self.today_max,NOW)

	def read_schedule(self,name):
		rest= timedelta(0)
		table = "schedule_raw"
		sql = "Select * from " + table + " where date between '" \
			+ str(self.today_min) + "' and '" + str(self.today_max - timedelta(seconds=1)) + "' and name='" + name + "' order by date"
		schdf = pd.read_sql_query(sql, self.__engine)
		global VALUE_OF_PLAN_ST,VALUE_OF_PLAN_ED
		if not schdf.empty:
			self.today_min=schdf.loc[schdf.raw_by=='morning','date'].min()
			self.today_max=schdf.loc[schdf.raw_by=='pre_last','date'].max()
			print(self.today_min,self.today_max)

		return schdf , rest


	# 2022/03/25 改用機台標準速度計算
	# class DB_connect.read_capacity 設變
	def read_capacity(self,name):
		table = "standard_speed"
		sql = "Select standard_CAP from " + table + " where name='"+ name + "'"
		workdf=pd.read_sql_query(sql, self.__engine)
		#print(workdf)
		# return workdf
		cap=workdf.loc[0,'standard_CAP'].astype(int)
		return cap

	def read_status(self,name):
		name=re.sub("\d+","", name)
		table = "status"
		sql = "Select value,status as '狀態' from " + table + " where name='"+ name + "'"
		seldf=pd.read_sql_query(sql, self.__engine)
		return seldf

		
	def read_from(self,name):
		t1=time.time()
		table = name

		# for test
		# seldf = pd.read_csv("testdb.csv", parse_dates=['date'])
		# return seldf

		#select ID between today
		sql = "Select id from " + table + " where date between '" + str(self.today_min) + "' and '" + str(self.today_max) + "'"
		iddf = pd.read_sql_query(sql, self.__engine)
		#print(iddf)
		if not iddf.empty and len(iddf) >=2 :
			idarr=tuple(iddf['id'].astype(int).values.tolist())
			#print(idarr)
			sql = "Select id,date,value,parts,work_order_id from " + table + " where id in "+ str(idarr)
			seldf = pd.read_sql_query(sql, self.__engine)
			#seldf=predf[predf['date'].between(today, now)]
			#print(seldf)
			print('read database cost %f sec' %(time.time()-t1))
			#return predf.iloc[0,0:].to_dict()
			return seldf
		else:
			seldf = pd.DataFrame({},columns=['id','date','value','parts','work_order_id'])
			return seldf

	def write_to_sql(self,df,table):
		sql = 'Select * from ' + table
		df.to_sql(table, self.__engine, if_exists='append', index=False)

	def close(self):
		self.__engine.dispose()		

if __name__ == '__main__':
	# dir=r'\\Nas\uta-share\UTA資料庫\UTA共用區\Q-專案執行\改善\(G組) MES\表單自動化' + '\\'
	dir='/home/uta_iot/excel_output/'
	if len(sys.argv) >1:
		NOW=datetime.strptime(sys.argv[1], "%Y%m%d").date()
		path=dir+sys.argv[1]+'_excel_output.xlsx'
	else:
		if datetime.today().weekday()==0 and datetime.today() < datetime.today().replace(hour=8):
			shift=timedelta(days=3)
		elif datetime.today() < datetime.today().replace(hour=8):
			shift=timedelta(days=1)
		else:
			shift=timedelta(days=0)
		NOW=datetime.today().date() - shift
		path=dir+str(NOW).replace('-','')+'_excel_output.xlsx'
	print(NOW)
	allst=time.time()
	# read from machine config
	machinedf=pd.read_excel("machine_config.xlsx")
	machinedf.columns=['name','host','parts','parts_type','status','status_type','UID','UID_type','WID','WID_type','option1','op_type1','option2','op_type2']
	machine = machinedf['name'].values.tolist()
	print(machine)
	print(time.time()-allst)

	oeedf=pd.DataFrame()
	piedf=pd.DataFrame()
	
	if os.path.exists(path):
		os.remove(path)
	wb = openpyxl.Workbook()
	# wb.remove_sheet(wb['Sheet'])
	del wb['Sheet']
	for i in range(len(machine)):
		name=machine[i].lower()	
		wb.create_sheet(title=name)
	wb.save(path)
	writer = pd.ExcelWriter(engine='openpyxl', path=path, mode='a',if_sheet_exists='replace')
	for i in range(len(machine)):
	# for i in range(1,2):
		conn = DB_connect()
		name=machine[i].lower()
		seldf = conn.read_from(name)
		stdf = conn.read_status(name)
		print(seldf)
		# if seldf is not None:
		if not seldf.empty:
			seldf.sort_values(by=['date'],inplace=True)
			# seldf.reset_index(drop=True,inplace=True)
			schdf, rest = conn.read_schedule(name)
			
			print(schdf)
			oee = OEE_Class(seldf, name,rest,schdf)
			#print(oee.workid)
			
			# 2022/03/25 cap設變
			# workdf = conn.read_capacity(oee.workid)
			workdf = conn.read_capacity(name)
			#print(workdf)
			t1=time.time()
			oee.calc_standrad(workdf)
			t2=time.time()
			print('calc standard cost %f sec' %(t2-t1))
			#print(oee.standard_pcs)
			oeedf = oee.calc_OEE()
			piedf = oee.piedf
			oeedf=oeedf.drop(columns=['nomal_min', 'nomal_max', 'nomal_avg', 'alarm_min', 'alarm_max',
       'alarm_avg'])
			mask=['Production', 'load_time', 'total_time',
        'ttr', 'mtbf', 'mttr']
			oeedf.loc[0,mask]=pd.to_timedelta(oeedf.loc[0,mask], unit='s').astype(str).str.replace('0 days ','')
			piedf['minutes']=round(piedf.during/60,1)
			tempS=pd.concat([pd.Series([name,NOW,500], index=['name','date','status']),piedf[['during','times','minutes']].loc[piedf['status']>=500].agg('sum')])
			if tempS.times>0:
				piedf=pd.concat([piedf[piedf['status']<500],tempS.to_frame(len(piedf)).T],axis=0)
			piedf['during']= pd.to_timedelta(piedf['during'], unit='s').astype(str)
			piedf.during=piedf.during.apply(lambda _:str(_).replace('0 days ',''))
			work_time = oee.work_time
			work_time[['during','MTBF','MTTR']]=work_time[['during','MTBF','MTTR']].apply(lambda _:_.astype(str).str.replace('0 days ',''))
			work_time[['MTBF','MTTR']]=work_time[['MTBF','MTTR']].apply(lambda _:_.astype(str).str.replace('NaT',''))
				
			last=oeedf.shape[0]+2
			oeedf.to_excel(writer,sheet_name=name)
			writer.if_sheet_exists='overlay'
			piedf=piedf.merge(stdf, left_on='status',right_on='value', how='left').drop(columns=['value'])
			sum=piedf[piedf.status<500].minutes.agg('sum')
			piedf=piedf.convert_dtypes(infer_objects=True)
			piedf['狀態']=piedf['狀態'].apply(lambda _:str(_).replace('<NA>','休息時間') if '<NA>' in str(_) else re.sub("\d+ ","", str(_)))
			piedf.status=piedf.status.astype(str)+"\n("+round((piedf.minutes/sum)*100,1).astype(str)+"%)\n"+ piedf['狀態'].astype(str)
			piedf['status']=piedf['status'].apply(lambda _:	"500+\n(0%)\n休息時間" if '500' in _ else _)
			piedf.to_excel(writer,sheet_name=name,startrow=last)
			
			last+=piedf.shape[0]+2
			work_time.to_excel(writer,sheet_name=name+"明細")

			ws= writer.sheets[name]
			ws_sub=writer.sheets[name+"明細"]
			for row in ws['A:R']:
				for cell in row:
					cell.font = Font(size=16)
			for j in range(1, ws.max_column+1):
				ws.column_dimensions[get_column_letter(j)].bestFit = True
				ws.column_dimensions[get_column_letter(j)].auto_size = True
				if j in (2,3,4,6,7,13):
					ws.column_dimensions[get_column_letter(j)].width = 16

			for j in range(1, ws_sub.max_column+1):
				ws_sub.column_dimensions[get_column_letter(j)].bestFit = True
				ws_sub.column_dimensions[get_column_letter(j)].auto_size = True
				if 2<=j<=3:
					ws_sub.column_dimensions[get_column_letter(j)].width = 23


			if piedf.shape[0]>0:
				last=oeedf.shape[0]+2
				min_row=last+1
				max_row=last+piedf.shape[0]+1
				for row in range(min_row+1,max_row+1):
					ws[f'F{row}'].number_format ='0次'
					ws[f'G{row}'].number_format ='0.0分'
				flag500=piedf.status.str.contains('500+').any()
				if flag500:
					max_row=max_row-1
				data1 = Reference(ws, min_col=7, min_row=min_row, max_col=7, max_row=max_row)
				data2 = Reference(ws, min_col=6, min_row=min_row, max_col=6, max_row=max_row)
				titles = Reference(ws, min_col=4, min_row=min_row+1, max_row=max_row)
				chart1 = BarChart()
				chart1.title = name + " 狀態分布"
				chart1.add_data(data=data1, titles_from_data=True)
				chart1.set_categories(titles)
				chart1.y_axis.title = '時間(分)'
				chart1.y_axis.numFmt='0"分"'
				# chart1.x_axis.title = '狀態'
				chart1.y_axis.majorUnit = 60
				chart1.y_axis.minorUnit = 60
				chart1.legend=None
				pp=ParagraphProperties(defRPr=CharacterProperties(sz=1400))
				rich_text = RichText(bodyPr=RichTextProperties(anchor="ctr",anchorCtr="1",rot="0",
					spcFirstLastPara="1",vertOverflow="ellipsis",wrap="square"),
					p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(sz=1400))])
				# chart1.x_axis.txPr = rich_text
				chart1.plot_area.dTable = DataTable()
				chart1.plot_area.dTable.showHorzBorder = True
				chart1.plot_area.dTable.showVertBorder = True
				chart1.plot_area.dTable.showOutline = True
				chart1.plot_area.dTable.showKeys = True
				chart1.plot_area.dTable.txPr = rich_text
				chart2 = LineChart()
				chart2.auto_axis = False
				chart2.add_data(data=data2, titles_from_data=True)
				chart2.series[0].marker.symbol = "circle"
				chart2.series[0].marker.size = 10
				chart2.y_axis.title = '次數'
				chart2.y_axis.axId = 200
				chart2.y_axis.crosses = "max"
				chart1.y_axis.txPr = rich_text
				chart2.y_axis.txPr = rich_text
				chart1.y_axis.title.tx.rich.p[0].pPr = pp
				chart2.y_axis.title.tx.rich.p[0].pPr = pp
				print(math.ceil(piedf.minutes.max()/60))
				# factor=round(piedf.minutes.max()/60,0)+1
				factor=math.ceil(piedf.minutes.max()/60)
				chart1.y_axis.scaling.max=factor*60
				if piedf.times.max()<=factor:
					chart2.y_axis.majorUnit = 1
					chart2.y_axis.scaling.max=factor
				elif piedf.times.max()<=factor*10:
					chart2.y_axis.majorUnit = 10
					chart2.y_axis.scaling.max=factor*10
				else:
					if factor!=0:
						n=math.ceil(piedf.times.max()/factor)
						chart2.y_axis.majorUnit = n
						chart2.y_axis.scaling.max = factor*n
				chart2.y_axis.scaling.min=0
				chart2.y_axis.minorUnit = 1
				chart1 += chart2
				chart1.style = 26
				# chart1.width = 22
				chart1.height = 15
				chart1.width = 45
				last=max_row+2 + (1 if flag500 else 0)
				ws.add_chart(chart1, "A"+str(last))
				ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
				ws.page_setup.paperSize = ws.PAPERSIZE_A3
				ws.sheet_properties.pageSetUpPr.fitToPage = True
				ws.print_options.headings=True
				ws_sub.page_setup.orientation = ws_sub.ORIENTATION_PORTRAIT
				ws_sub.page_setup.paperSize = ws_sub.PAPERSIZE_A4
				ws_sub.print_title_rows='1:1'
				ws_sub.sheet_properties.pageSetUpPr.fitToPage = True
				ws_sub.page_setup.fitToHeight = False
				ws_sub.print_options.gridLines=True
				ws_sub.oddHeader.center.text = name + "明細"
				# ws_sub.freeze_panes = 'A2'
				# ws_sub.print_options.headings=True


			print(work_time)
			print(oeedf)
			print(piedf)


	
	writer.save()

	# wb.save(path)
	# wb.close()

	
	# workbook = Workbook(path)
	# Create and set PDF options
	# pdfOptions = PdfSaveOptions()
	# workbook.save(path.replace("_excel_output.xlsx",".pdf"), pdfOptions)
	conn.close()
	alled = time.time()
	# 列印結果
	print("loop cost %f sec" % (alled - allst))  # 會自動做進位
	



